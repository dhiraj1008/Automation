import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(file):
    wb = xl.load_workbook(file)
    sheet = wb['Sheet1']

    for i in range(2, sheet.max_row + 1):
        cell = sheet.cell(i, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(i, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_col=4, min_row=2, max_col=4, max_row=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(file)

